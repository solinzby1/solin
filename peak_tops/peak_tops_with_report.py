import subprocess
import openpyxl
from openpyxl import load_workbook
from statistics import geometric_mean
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import os
import re


output_report = "gemm_peak_test.xlsx"

commands = [
    "bash gemm_peak_test.sh > gemm_peak_test.log",
]

def create_new_workbook(file_path):
    # Delete the existing file if it exists
    if os.path.exists(file_path):
        os.remove(file_path)

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    workbook.save(file_path)

def execute_shell_commands(commands):
    for command in commands:
        print(f"Executing: {command}")
        subprocess.run(command, shell=True)

def extract_and_store_gemm_hipblaslt_mnk(log_file_path, excel_file_path, worksheet_name):
    with open(log_file_path, 'r') as file:
        lines = file.readlines()

    data = []
    current_datatype_col1 = None
    current_datatype_col2 = None
    for line in lines:
        line = line.strip()
        if line in ['fp16', 'bf16','fp64', 'fp8', 'fp32', 'tf32', 'int8']:
            current_datatype_col1 = line
        elif line in ['0', '1', '2', '3', '4', '5', '6', '7']:
            current_datatype_col2 = line
        else:
            if current_datatype_col1 is not None and current_datatype_col2 is not None:
                 gflops, latency = map(float, line.split(','))
                 data.append([current_datatype_col1, current_datatype_col2, gflops / 1000, latency])
    df = pd.DataFrame(data, columns=['datatype', 'GPU', 'Tflops', 'latency (us)'])

    with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a') as writer:
        if worksheet_name not in writer.book.sheetnames:
            df.to_excel(writer, sheet_name=worksheet_name, index=False)
        else:
            startrow = writer.book[worksheet_name].max_row
            df.to_excel(writer, sheet_name=worksheet_name, index=False, startrow=startrow + 1, header=False)


# Call the function to create a new workbook with the specified file path
create_new_workbook(output_report)

# Call the function to execute the shell commands
execute_shell_commands(commands)

print("analysis and save log... ")

extract_and_store_gemm_hipblaslt_mnk("gemm_peak_test.log", output_report, "gemm_peak")

print("anlysis end")
